#!/usr/bin/python
# -*- coding: UTF-8 -*-


import requests
import json
import datetime
import time
import os
import sys
import subprocess

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

from errno import EEXIST
from os import makedirs, path

reload(sys)
sys.setdefaultencoding('utf-8')

script_dir = os.path.dirname(__file__)
fileName = sys.argv[0][sys.argv[0].rfind(os.sep) + 1:].rstrip('.py')

imgs_dir = os.path.join(script_dir, fileName + '/imgs/')
files_dir = os.path.join(script_dir, fileName + '/files/')

allTableName = []
allBucketsDict_keyTableName = {}
allDataDict_keyErrorCode = {}


def setupAllTableName(startDate, endDate):
    datestart = datetime.datetime.strptime(startDate, '%Y-%m-%d')
    dateend = datetime.datetime.strptime(endDate, '%Y-%m-%d')

    while datestart <= dateend:
        erveryDay = datestart.strftime('app-md-%Y-%m-%d')
        allTableName.append(erveryDay)
        datestart += datetime.timedelta(days=1)


def mkdir_p(mypath):
    '''Creates a directory. equivalent to using mkdir -p on the command line'''
    try:
        makedirs(mypath)
    except OSError as exc:  # Python >2.5
        if exc.errno == EEXIST and path.isdir(mypath):
            pass
        else:
            raise


def removeFileAtPath(path):
    if os.path.exists(path):
        # 删除文件，可使用以下两种方法。
        os.remove(path)


def requestDataFromNet_AllTable(sql):
    for x in xrange(0, len(allTableName)):
        tableName = allTableName[x]

        try:
            url = 'http://192.168.223.192/essql_sql'
            headers = {'Content-Type': 'application/json;charset=utf-8'}
            r = requests.get(url, headers=headers, data=sql)
            json = r.json()
            return json
        except Exception, e:
            raise e
        finally:
            pass


def covertDateStringFromAllTableName():
    allDateStringAry = []
    for x in xrange(0, len(allTableName)):
        tableName = allTableName[x]
        dateString = tableName.replace('app-md-', '')
        allDateStringAry.append(dateString)
    return allDateStringAry


def creatNewExcel():
    # 使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
    wb = Workbook()
    # 一个工作簿(workbook)在创建的时候同时至少也新建了一张工作表(worksheet)。
    # 你可以通过openpyxl.workbook.Workbook.active()调用得到正在运行的工作表。
    ws = wb.active

    # 使用openpyxl.workbook.Workbook.create_sheet()新建一张表
    ws1 = wb.create_sheet()  # 默认插在工作簿末尾
    ws2 = wb.create_sheet(0)  # 插入在工作簿的第一个位置
    ws.title = "New Title"  # 系统第一张表默认的名称Sheet
    return wb, ws


def saveExcel(wb, ws, excelFilePath):
    excelAdjustColumn(ws)
    # 特别警告：这个操作将会在没有认识提示的情况下用现在写的内容，覆盖掉原文件中的所有内容
    wb.save(excelFilePath)
    subprocess.call(["open", "-R", excelFilePath])


def excelAdjustColumn(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)))) + 0.5

    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def saveDataToDict(tableName, buckets):
    allBucketsDict_keyTableName[tableName] = buckets

    eachDict = {}
    for bucket in buckets:
        key = bucket['key']
        # if key == 'A connection failure occurred: SSL problem (Possible causes may include a bad/expired/self-signed certificate, clock set to wrong date)':
        # 	print bucket['doc_count']
        # print key
        eachDict[key] = bucket
    allDataDict_keyErrorCode[tableName] = eachDict


def requestEveryErrorCount(moduleName, platForm):
    for x in xrange(0, len(allTableName)):
        tableName = allTableName[x]

        # sql = 'SELECT codelogsubtype,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        # if len(moduleName)>0:
        # 	sql = sql + '\nand userdefined.rn_module_name = "'+moduleName+'"'

        # sql = sql + '\nand codelogsubtype in '+\
        # '("147104","147001","147002","147003","147004","147005","147006","147007")'

        # sql = sql + '\nGROUP BY codelogsubtype'

        # sql = 'SELECT userdefined.responseStatusCode,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand userdefined.responseStatusCode="404"'

        # sql = sql + '\nGROUP BY userdefined.responseStatusCode'

        # sql = 'SELECT userdefined.error,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        # if len(moduleName)>0:
        # 	sql = sql + '\nand userdefined.rn_module_name = "'+moduleName+'"'

        # sql = sql + '\nand codelogsubtype = 147001'

        # sql = sql + '\nGROUP BY userdefined.error'

        # sql = 'SELECT userdefined.rn_module_name,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        # sql = sql + '\nand userdefined.realUrl like "%/download/plugins/static%"'

        # sql = sql + '\nGROUP BY userdefined.rn_module_name'

        sql = 'SELECT codelogsubtype,count(*) \nFROM ' + tableName + ' \nwhere deviceostype = "'
        if len(platForm) > 0:
            sql = sql + platForm + '"'
        else:
            sql = sql + 'iOS' + '"'

        sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        # sql = sql + '\nand userdefined.realUrl like "%/download/plugins/static%"'

        if len(moduleName) > 0:
            sql = sql + '\nand userdefined.rn_module_name = "' + moduleName + '"'

        sql = sql + '\nGROUP BY codelogsubtype'

        # sql = 'SELECT userdefined.responseStatusCode,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        # sql = sql + '\nand userdefined.realUrl like "%/download/plugins/static%"'

        # if len(moduleName)>0:
        # 	sql = sql + '\nand userdefined.rn_module_name = "'+moduleName+'"'

        # sql = sql + '\nGROUP BY userdefined.responseStatusCode'

        # sql = 'SELECT userdefined.error,count(*) \nFROM '+tableName+' \nwhere deviceostype = "'
        # if len(platForm)>0:
        # 	sql = sql + platForm + '"'
        # else:
        # 	sql = sql + 'iOS' + '"'

        # sql = sql + '\nand appversion in ("9.6.5","9.6.0")'
        # sql = sql + '\nand codelogsubtype="147001"'
        # sql = sql + '\nand userdefined.error in ("A connection failure occurred", "The request timed out","","A connection failure occurred: SSL problem (Possible causes may include a bad/expired/self-signed certificate, clock set to wrong date)")'

        # if len(moduleName)>0:
        # 	sql = sql + '\nand userdefined.rn_module_name = "'+moduleName+'"'

        # # sql = sql + '\nand userdefined.realUrl like "%/download/plugins/static%"'

        # sql = sql + '\nGROUP BY userdefined.error'

        json = requestDataFromNet_AllTable(sql)

        # buckets = json["aggregations"]["userdefined.error"]["buckets"]
        buckets = json["aggregations"]["codelogsubtype"]["buckets"]

        saveDataToDict(tableName, buckets)


    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'')
    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_Car_SelectCar_allCode_Disaster_count.xlsx')
    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_YoungPlugin_allCode_count.xlsx')
    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_allCode_Disaster_count.xlsx')
    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_Models_Disaster_count.xlsx')
    # drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_Disaster_EveryError_count.xlsx')
    drawExcelEveryErrorCount(moduleName,platForm,allTableName,'EveryDay_147001_copy_count.xlsx')

def requestSingleErrorCount(moduleName, platForm):
    for x in xrange(0, len(allTableName)):
        tableName = allTableName[x]

        sql = 'SELECT * \nFROM ' + tableName + ' \nwhere deviceostype = "'
        if len(platForm) > 0:
            sql = sql + platForm + '"'
        else:
            sql = sql + 'iOS' + '"'

        sql = sql + '\nand appversion in ("9.6.5","9.6.0")'

        if len(moduleName) > 0:
            sql = sql + '\nand userdefined.rn_module_name = "' + moduleName + '"'

        sql = sql + '\nand codelogsubtype = 147001'

        sql = sql + '\nand userdefined.error like "%Failed to move file from%"'

        json = requestDataFromNet_AllTable(sql)

        totalCount = json["hits"]["total"]

        bucket = {}
        bucket['doc_count'] = totalCount

        bucketDict = {}
        bucketDict['Failed  to move file from  XXX  to  XXX'] = bucket

        allDataDict_keyErrorCode[tableName] = bucketDict

    drawExcelEveryErrorCount(moduleName, platForm, allTableName, '')


def drawExcelEveryErrorCount(moduleName, platForm, columnNames, fileSavePath):
    excelFilePath = files_dir
    if len(moduleName) > 0:
        excelFilePath = excelFilePath + moduleName + '_'

    if len(platForm) > 0:
        excelFilePath = excelFilePath + platForm + '_'

    excelFilePath = excelFilePath + 'EveryDay_EveryError_count.xlsx'

    if len(fileSavePath) > 0:
        excelFilePath = fileSavePath

    (wb, ws) = creatNewExcel()

    # column_ones = ["147104","147001","147002","147003","147004","147005","147006","147007"]
    column_ones = ["147001", "147002", "147003", "147004", "147005", "147006", "147007"]
    # column_ones = ["404"]
    # column_ones = ["Failed  to move file from  XXX  to  XXX"]

    # column_ones = ["404","0","200","503","403","416"]

    # column_ones = ["Car_SelectCar","YoungPlugin","DianPing_Main","SevenStepsBuyCar_Step2","SevenStepsBuyCar_Step0","CarService_Violation","Car_SeriesSummary","Club_OwnerPrice","UsedCar_Search","Car_CalculateRN","SevenStepsBuyCar_Step1","SevenStepsBuyCar_Step3","UsedCar_Detail"]

    # column_ones = ["The request timed out","A connection failure occurred","A connection failure occurred: SSL problem (Possible causes may include a bad/expired/self-signed certificate, clock set to wrong date)",""]

    for x in xrange(0, len(column_ones)):
        column_one = column_ones[x]
        vcell = ws.cell(row=x + 2, column=1)
        vcell.value = column_one

    for x in xrange(0, len(columnNames)):
        tableName = columnNames[x]
        dateString = tableName.replace('app-md-', '')
        dateCell = ws.cell(row=1, column=x + 2)
        dateCell.value = dateString

        bucketDict = allDataDict_keyErrorCode[tableName]

        for j in xrange(0, len(column_ones)):
            rowKey = column_ones[j]
            cell = ws.cell(row=j + 2, column=x + 2)
            try:
                bucket = bucketDict[rowKey]

            except Exception, e:
                cell.value = 0
                pass
            else:
                if rowKey == '147007':
                    # 147007成功类型没有全部上报，乘10会接近真实数据
                    cell.value = bucket['doc_count'] * 10
                else:
                    cell.value = bucket['doc_count']

    saveExcel(wb, ws, excelFilePath)


def getDataEveryHour(date, span, moduleName, platForm):
    completeDate = date + ' 00:00:00'
    convertTimeStamp = time.mktime(time.strptime(completeDate, '%Y-%m-%d %H:%M:%S'))

    startTimeStamp = convertTimeStamp
    endTimeStamp = convertTimeStamp
    onrHoure = 3600

    allTimeAry = []
    for x in xrange(0, span * 24):
        startTimeStamp = endTimeStamp
        endTimeStamp = startTimeStamp + onrHoure

        startTimeStampString = time.localtime(startTimeStamp)
        endTimeStampString = time.localtime(endTimeStamp)

        startDateString = time.strftime('%Y-%m-%d %H:%M:%S', startTimeStampString)
        endDateString = time.strftime('%Y-%m-%d %H:%M:%S', endTimeStampString)

        queryDate = startDateString[0:10]

        sql = 'SELECT codelogsubtype,count(*) ' + \
              '\nFROM app-md-' + queryDate + \
              ' \nwhere deviceostype="' + platForm + '"' + \
              ' \nand userdefined.rn_module_name="' + moduleName + \
              '"\nand writelogdbtime between ' + \
              str(startTimeStamp * 1000) + \
              ' and ' + str(endTimeStamp * 1000) + \
              '\nGROUP BY codelogsubtype'

        json = requestDataFromNet_AllTable(sql)

        buckets = json["aggregations"]["codelogsubtype"]["buckets"]

        columnKey = startDateString + '-' + endDateString
        allTimeAry.append(columnKey)
        saveDataToDict(columnKey, buckets)

    drawExcelEveryErrorCount(moduleName, platForm, allTimeAry,
                             date + "_" + moduleName + "_" + platForm + "_" + "ErrorCodeCount.xlsx")

def loadSQLData(path):
    excelFilePath = files_dir + 'patch_SQL.xlsx'
    (wb, ws) = creatNewExcel()
    try:
        f = open(path, 'r')  # 打开文件
        lineNum = 0
        while True:
            lineNum += 1
            line = f.readline()  # 逐行读取
            if not line:
                break
            lineAry = line.split(', ')
            for x in xrange(0, len(lineAry)):
                vcell = ws.cell(row=lineNum, column=x+1)
                vcell.value = lineAry[x]

    finally:
        if f:
            f.close()  # 确保文件被关闭
        saveExcel(wb, ws, excelFilePath)

def main():
    mkdir_p(imgs_dir)
    mkdir_p(files_dir)

    loadSQLData(files_dir+'/patch.txt')

    # setupAllTableName('2018-11-05', '2018-12-02')

    # allDateStringAry = covertDateStringFromAllTableName()

    # requestEveryErrorCount('', 'iOS')


# requestEveryErrorCount('DianPing_Main','iOS')
# requestEveryErrorCount('YoungPlugin','iOS')
# requestEveryErrorCount('','iOS')
# requestSingleErrorCount('','iOS')

# getDataEveryHour('2018-11-02',2,'SevenStepsBuyCar_Entry','Android')

if __name__ == "__main__":
    main()